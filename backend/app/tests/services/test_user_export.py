from app.models.user import User
from app.services.user_export import _cell, _is_money_header, build_workbook
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from tests.utils import create_test_account, create_test_expense, create_test_user


async def test_workbook_only_contains_owner_rows(
    db_session: AsyncSession, test_user: User
):
    account = await create_test_account(db_session, owner_id=test_user.id, name="Mine")
    await create_test_expense(
        db_session,
        owner_id=test_user.id,
        account_id=account.id,
        amount=100.1031349132,
        description="owner lunch",
    )

    other = await create_test_user(db_session, email="export-other@example.com")
    other_account = await create_test_account(db_session, owner_id=other.id, name="Theirs")
    await create_test_expense(
        db_session,
        owner_id=other.id,
        account_id=other_account.id,
        description="secret expense",
    )

    path, counts = await build_workbook(db_session, test_user)
    try:
        wb = load_workbook(path)
        assert "items" not in wb.sheetnames
        assert "expenses" in wb.sheetnames
        assert "_manifest" in wb.sheetnames

        profile = wb["profile"]
        headers = [cell.value for cell in next(profile.iter_rows(min_row=1, max_row=1))]
        assert "hashed_password" not in headers
        assert "email" in headers

        expense_sheet = wb["expenses"]
        expense_headers = [
            cell.value for cell in next(expense_sheet.iter_rows(min_row=1, max_row=1))
        ]
        desc_idx = expense_headers.index("description")
        amount_idx = expense_headers.index("amount")
        descriptions = []
        amounts = []
        for row in expense_sheet.iter_rows(min_row=2, values_only=True):
            if row[desc_idx]:
                descriptions.append(row[desc_idx])
            amounts.append(row[amount_idx])
        assert "owner lunch" in descriptions
        assert "secret expense" not in descriptions
        assert counts["expenses"] == 1
        assert amounts == [100.10]
    finally:
        path.unlink(missing_ok=True)


async def test_empty_user_still_gets_workbook(db_session: AsyncSession, test_user: User):
    path, counts = await build_workbook(db_session, test_user)
    try:
        wb = load_workbook(path)
        assert counts["profile"] == 1
        assert counts["expenses"] == 0
        assert set(wb.sheetnames) >= {
            "profile",
            "accounts",
            "expenses",
            "incomes",
            "_manifest",
        }
    finally:
        path.unlink(missing_ok=True)


def test_cash_balance_flag_is_not_rounded_as_money():
    assert _is_money_header("affects_cash_balance") is False
    assert _cell(True, header="affects_cash_balance") is True
    assert _cell(False, header="affects_cash_balance") is False
    assert _cell(100.1031349132, header="amount") == 100.10
    assert _cell(100.1031349132, header="total_amount") == 100.10
    assert _cell(1.239, header="fees") == 1.24
