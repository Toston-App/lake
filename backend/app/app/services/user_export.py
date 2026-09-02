from __future__ import annotations

import hashlib
import logging
import os
import tempfile
from collections.abc import Awaitable, Callable
from datetime import date, datetime, timezone
from decimal import ROUND_HALF_UP, Decimal
from enum import Enum
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app import crud
from app.core.config import settings
from app.models.account import Account
from app.models.asset import Asset
from app.models.balance_adjustment import BalanceAdjustment
from app.models.category import Category
from app.models.expense import Expense
from app.models.holding import Holding
from app.models.income import Income
from app.models.investment_transaction import InvestmentTransaction
from app.models.place import Place
from app.models.subcategory import Subcategory
from app.models.transfer import Transfer
from app.models.user import User
from app.services import r2

logger = logging.getLogger(__name__)

SCHEMA_VERSION = 1
EXCEL_MAX_ROWS = 1_048_576
GENERIC_FAIL = "Export failed. Please try again."

ProgressCb = Callable[[int, str], Awaitable[None]]

_MONEY_TOKENS = (
    "amount",
    "balance",
    "total",
    "value",
    "price",
    "fee",
    "cost",
    "invested",
    "pct",
)
_NOT_MONEY_HEADERS = {
    "quantity",
    "affects_cash_balance",
    "cost_currency",
}


def _is_money_header(name: str) -> bool:
    if name in _NOT_MONEY_HEADERS:
        return False
    lowered = name.lower()
    return any(token in lowered for token in _MONEY_TOKENS)


def _as_money(value: Any) -> float:
    quantized = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return float(quantized)


def _cell(value: Any, *, header: str | None = None) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, Enum):
        return value.value
    if header and _is_money_header(header) and isinstance(value, (int, float, Decimal)):
        return _as_money(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, UUID):
        return str(value)
    return value


def _excel_value(ws, header: str, value: Any) -> Any:
    converted = _cell(value, header=header)
    if converted is None or not _is_money_header(header):
        return converted
    cell = WriteOnlyCell(ws, value=converted)
    cell.number_format = "0.00"
    return cell


def export_filename(when: datetime | None = None) -> str:
    when = when or datetime.now(timezone.utc)
    return f"toston-export-{when.date().isoformat()}.xlsx"


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


async def _stream_rows(db: AsyncSession, stmt):
    result = await db.stream(stmt)
    async for row in result:
        yield row


async def _write_query_sheet(
    db: AsyncSession,
    ws,
    stmt,
    headers: list[str],
    *,
    base_name: str,
    workbook: Workbook,
) -> int:
    ws.append(headers)
    count = 0
    sheet_rows = 1
    current = ws
    suffix = 2
    async for row in _stream_rows(db, stmt):
        mapping = row._mapping if hasattr(row, "_mapping") else None
        if mapping is not None:
            values = [_excel_value(current, h, mapping[h]) for h in headers]
        else:
            values = [_excel_value(current, h, v) for h, v in zip(headers, row, strict=False)]
        if sheet_rows >= EXCEL_MAX_ROWS:
            current = workbook.create_sheet(f"{base_name}_{suffix}"[:31])
            current.append(headers)
            sheet_rows = 1
            suffix += 1
        current.append(values)
        sheet_rows += 1
        count += 1
    return count


async def build_workbook(
    db: AsyncSession,
    user: User,
    on_progress: ProgressCb | None = None,
) -> tuple[Path, dict[str, int]]:
    wb = Workbook(write_only=True)
    counts: dict[str, int] = {}
    weights = {
        "profile": 2,
        "accounts": 5,
        "categories": 5,
        "subcategories": 5,
        "places": 5,
        "expenses": 30,
        "incomes": 20,
        "transfers": 8,
        "imports": 3,
        "feedback": 2,
        "balance_adjustments": 5,
        "holdings": 4,
        "investment_transactions": 5,
        "assets": 3,
        "_manifest": 1,
    }
    total_w = sum(weights.values())
    done_w = 0

    async def step(name: str) -> None:
        nonlocal done_w
        done_w += weights[name]
        if on_progress:
            await on_progress(int(done_w * 100 / total_w), name)

    # profile
    ws = wb.create_sheet("profile")
    ws.append(
        [
            "email",
            "balance_total",
            "balance_income",
            "balance_outcome",
            "created_at",
        ]
    )
    ws.append(
        [
            _excel_value(ws, "email", user.email),
            _excel_value(ws, "balance_total", user.balance_total),
            _excel_value(ws, "balance_income", user.balance_income),
            _excel_value(ws, "balance_outcome", user.balance_outcome),
            _excel_value(ws, "created_at", user.created_at),
        ]
    )
    counts["profile"] = 1
    await step("profile")

    owner = user.id

    ws = wb.create_sheet("accounts")
    counts["accounts"] = await _write_query_sheet(
        db,
        ws,
        select(
            Account.name,
            Account.type,
            Account.initial_balance,
            Account.current_balance,
            Account.total_expenses,
            Account.total_incomes,
            Account.total_transfers_in,
            Account.total_transfers_out,
            Account.total_investments_usd,
            Account.total_investments_mxn,
            Account.created_at,
        )
        .where(Account.owner_id == owner)
        .order_by(Account.id),
        [
            "name",
            "type",
            "initial_balance",
            "current_balance",
            "total_expenses",
            "total_incomes",
            "total_transfers_in",
            "total_transfers_out",
            "total_investments_usd",
            "total_investments_mxn",
            "created_at",
        ],
        base_name="accounts",
        workbook=wb,
    )
    await step("accounts")

    ws = wb.create_sheet("categories")
    counts["categories"] = await _write_query_sheet(
        db,
        ws,
        select(
            Category.name,
            Category.description,
            Category.is_income,
            Category.total,
            Category.created_at,
        )
        .where(Category.owner_id == owner)
        .order_by(Category.id),
        [
            "name",
            "description",
            "is_income",
            "total",
            "created_at",
        ],
        base_name="categories",
        workbook=wb,
    )
    await step("categories")

    ws = wb.create_sheet("subcategories")
    counts["subcategories"] = await _write_query_sheet(
        db,
        ws,
        select(
            Subcategory.name,
            Subcategory.description,
            Category.name.label("category_name"),
            Subcategory.total,
            Subcategory.created_at,
        )
        .outerjoin(Category, Category.id == Subcategory.category_id)
        .where(Subcategory.owner_id == owner)
        .order_by(Subcategory.id),
        [
            "name",
            "description",
            "category_name",
            "total",
            "created_at",
        ],
        base_name="subcategories",
        workbook=wb,
    )
    await step("subcategories")

    ws = wb.create_sheet("places")
    counts["places"] = await _write_query_sheet(
        db,
        ws,
        select(
            Place.name,
            Place.is_online,
            Place.created_at,
        )
        .where(Place.owner_id == owner)
        .order_by(Place.id),
        ["name", "is_online", "created_at"],
        base_name="places",
        workbook=wb,
    )
    await step("places")

    ws = wb.create_sheet("expenses")
    counts["expenses"] = await _write_query_sheet(
        db,
        ws,
        select(
            Expense.amount,
            Expense.date,
            Expense.description,
            Account.name.label("account_name"),
            Category.name.label("category_name"),
            Subcategory.name.label("subcategory_name"),
            Place.name.label("place_name"),
            Expense.made_from,
            Expense.created_at,
        )
        .outerjoin(Account, Account.id == Expense.account_id)
        .outerjoin(Category, Category.id == Expense.category_id)
        .outerjoin(Subcategory, Subcategory.id == Expense.subcategory_id)
        .outerjoin(Place, Place.id == Expense.place_id)
        .where(Expense.owner_id == owner)
        .order_by(Expense.id),
        [
            "amount",
            "date",
            "description",
            "account_name",
            "category_name",
            "subcategory_name",
            "place_name",
            "made_from",
            "created_at",
        ],
        base_name="expenses",
        workbook=wb,
    )
    await step("expenses")

    ws = wb.create_sheet("incomes")
    counts["incomes"] = await _write_query_sheet(
        db,
        ws,
        select(
            Income.amount,
            Income.date,
            Income.description,
            Account.name.label("account_name"),
            Subcategory.name.label("subcategory_name"),
            Place.name.label("place_name"),
            Income.made_from,
            Income.created_at,
        )
        .outerjoin(Account, Account.id == Income.account_id)
        .outerjoin(Subcategory, Subcategory.id == Income.subcategory_id)
        .outerjoin(Place, Place.id == Income.place_id)
        .where(Income.owner_id == owner)
        .order_by(Income.id),
        [
            "amount",
            "date",
            "description",
            "account_name",
            "subcategory_name",
            "place_name",
            "made_from",
            "created_at",
        ],
        base_name="incomes",
        workbook=wb,
    )
    await step("incomes")

    account_from = aliased(Account)
    account_to = aliased(Account)
    ws = wb.create_sheet("transfers")
    counts["transfers"] = await _write_query_sheet(
        db,
        ws,
        select(
            Transfer.amount,
            Transfer.date,
            Transfer.description,
            account_from.name.label("from_account_name"),
            account_to.name.label("to_account_name"),
            Transfer.created_at,
        )
        .outerjoin(account_from, account_from.id == Transfer.from_acc)
        .outerjoin(account_to, account_to.id == Transfer.to_acc)
        .where(Transfer.owner_id == owner)
        .order_by(Transfer.id),
        [
            "amount",
            "date",
            "description",
            "from_account_name",
            "to_account_name",
            "created_at",
        ],
        base_name="transfers",
        workbook=wb,
    )
    await step("transfers")

    ws = wb.create_sheet("balance_adjustments")
    counts["balance_adjustments"] = await _write_query_sheet(
        db,
        ws,
        select(
            Account.name.label("account_name"),
            BalanceAdjustment.old_balance,
            BalanceAdjustment.new_balance,
            BalanceAdjustment.adjustment_amount,
            BalanceAdjustment.description,
            BalanceAdjustment.adjustment_date,
            BalanceAdjustment.created_at,
        )
        .outerjoin(Account, Account.id == BalanceAdjustment.account_id)
        .where(BalanceAdjustment.owner_id == owner)
        .order_by(BalanceAdjustment.id),
        [
            "account_name",
            "old_balance",
            "new_balance",
            "adjustment_amount",
            "description",
            "adjustment_date",
            "created_at",
        ],
        base_name="balance_adjustments",
        workbook=wb,
    )
    await step("balance_adjustments")

    ws = wb.create_sheet("holdings")
    counts["holdings"] = await _write_query_sheet(
        db,
        ws,
        select(
            Account.name.label("account_name"),
            Asset.symbol.label("asset_symbol"),
            Asset.name.label("asset_name"),
            Holding.quantity,
            Holding.avg_cost_basis,
            Holding.cost_currency,
            Holding.total_invested,
            Holding.current_value,
            Holding.current_value_usd,
            Holding.current_value_mxn,
            Holding.unrealized_gain_loss,
            Holding.unrealized_gain_loss_pct,
            Holding.created_at,
        )
        .outerjoin(Account, Account.id == Holding.account_id)
        .outerjoin(Asset, Asset.id == Holding.asset_id)
        .where(Holding.owner_id == owner)
        .order_by(Holding.id),
        [
            "account_name",
            "asset_symbol",
            "asset_name",
            "quantity",
            "avg_cost_basis",
            "cost_currency",
            "total_invested",
            "current_value",
            "current_value_usd",
            "current_value_mxn",
            "unrealized_gain_loss",
            "unrealized_gain_loss_pct",
            "created_at",
        ],
        base_name="holdings",
        workbook=wb,
    )
    await step("holdings")

    ws = wb.create_sheet("investment_transactions")
    counts["investment_transactions"] = await _write_query_sheet(
        db,
        ws,
        select(
            Account.name.label("account_name"),
            InvestmentTransaction.transaction_type,
            InvestmentTransaction.quantity,
            InvestmentTransaction.price_per_unit,
            InvestmentTransaction.currency,
            InvestmentTransaction.total_amount,
            InvestmentTransaction.fees,
            InvestmentTransaction.affects_cash_balance,
            InvestmentTransaction.notes,
            InvestmentTransaction.executed_at,
            InvestmentTransaction.created_at,
        )
        .outerjoin(Account, Account.id == InvestmentTransaction.account_id)
        .where(InvestmentTransaction.owner_id == owner)
        .order_by(InvestmentTransaction.id),
        [
            "account_name",
            "transaction_type",
            "quantity",
            "price_per_unit",
            "currency",
            "total_amount",
            "fees",
            "affects_cash_balance",
            "notes",
            "executed_at",
            "created_at",
        ],
        base_name="investment_transactions",
        workbook=wb,
    )
    await step("investment_transactions")

    ws = wb.create_sheet("assets")
    counts["assets"] = await _write_query_sheet(
        db,
        ws,
        select(
            Asset.symbol,
            Asset.name,
            Asset.asset_class,
            Asset.asset_type,
            Asset.currency,
            Asset.market,
            Asset.sector,
            Asset.country,
        )
        .where(
            Asset.id.in_(select(Holding.asset_id).where(Holding.owner_id == owner))
        )
        .order_by(Asset.id),
        [
            "symbol",
            "name",
            "asset_class",
            "asset_type",
            "currency",
            "market",
            "sector",
            "country",
        ],
        base_name="assets",
        workbook=wb,
    )
    await step("assets")

    ws = wb.create_sheet("_manifest")
    ws.append(["key", "value"])
    ws.append(["schema_version", SCHEMA_VERSION])
    ws.append(["exported_at", datetime.now(timezone.utc).isoformat()])
    ws.append(["user_id", user.id])
    for name, n in counts.items():
        ws.append([f"rows_{name}", n])
    counts["_manifest"] = 1
    await step("_manifest")

    fd, tmp_name = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    path = Path(tmp_name)
    try:
        wb.save(path)
    except Exception:
        path.unlink(missing_ok=True)
        raise
    return path, counts


async def process_export_job(db: AsyncSession, export_id: UUID) -> None:
    export = await crud.data_export.get(db, export_id)
    if export is None:
        return
    user = await crud.user.get(db, id=export.owner_id)
    if user is None:
        await crud.data_export.mark_failed(
            db, export, error="User not found for this export."
        )
        return

    path: Path | None = None
    uploaded_key: str | None = None
    try:

        async def on_progress(pct: int, step: str) -> None:
            current = await crud.data_export.get(db, export_id)
            if current is None:
                return
            await crud.data_export.mark_progress(db, current, pct=pct, step=step)

        path, _counts = await build_workbook(db, user, on_progress)
        key = r2.object_key(user, export.id)
        await r2.upload_file(
            key, path, user_uuid=user.uuid, export_id=export.id
        )
        uploaded_key = key
        sha = _sha256_file(path)
        size = path.stat().st_size
        current = await crud.data_export.get(db, export_id)
        if current is None:
            return
        await crud.data_export.mark_ready(
            db,
            current,
            object_key=key,
            content_sha256=sha,
            byte_size=size,
            retention_days=settings.EXPORT_RETENTION_DAYS,
        )
    except Exception:
        logger.exception("Export job %s failed", export_id)
        if uploaded_key:
            try:
                await r2.delete_object(uploaded_key)
            except Exception:
                logger.exception("Failed to delete partial export object")
        current = await crud.data_export.get(db, export_id)
        if current is not None:
            await crud.data_export.mark_failed(db, current, error=GENERIC_FAIL)
    finally:
        if path is not None:
            path.unlink(missing_ok=True)
